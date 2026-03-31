#!/usr/bin/env python3
"""
Публикация dashboard_output.xlsx в Google Sheets с полным форматированием.

Предварительная настройка (один раз):
  1. Открой https://console.cloud.google.com/
  2. Создай проект → API & Services → Enable APIs → включи Google Sheets API и Google Drive API
  3. Credentials → Create Credentials → OAuth 2.0 Client ID → Desktop App
  4. Скачай JSON → сохрани как ~/.config/homeart_oauth.json
  5. Запусти этот скрипт — в браузере откроется окно авторизации (один раз)
  6. После этого токен сохраняется, авторизация больше не нужна

Использование:
  python3 sheets_upload.py dashboard_output.xlsx           # создаёт новую таблицу
  python3 sheets_upload.py dashboard_output.xlsx <ID>      # обновляет существующую
"""

import sys
import os
import warnings
warnings.filterwarnings("ignore")

import openpyxl
from openpyxl.styles import PatternFill

CREDENTIALS_FILE = os.path.expanduser("~/.config/homeart_oauth.json")
TOKEN_FILE       = os.path.expanduser("~/.config/homeart_token.json")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# Цвета (hex без #)
COLOR_HEADER      = "2F5496"   # синий заголовок
COLOR_HEADER_TEXT = "FFFFFF"   # белый текст
COLOR_HIGHLIGHT   = "E2EFDA"   # зелёный highlight
COLOR_ALT_ROW     = "F2F2F2"   # серый чётные строки
COLOR_TOTAL       = "D6E4F0"   # голубой итоги
COLOR_LOST        = "FCE4EC"   # красный — не реализовано
COLOR_WON         = "E8F5E9"   # зелёный — успешно
COLOR_NOTE        = "FFF9C4"   # жёлтый — примечание


# ── Auth ──────────────────────────────────────────────────────────────────────

def get_creds():
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request

    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists(CREDENTIALS_FILE):
                print(f"\n❌ Файл учётных данных не найден: {CREDENTIALS_FILE}")
                sys.exit(1)
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, "w") as tok:
            tok.write(creds.to_json())
    return creds


# ── Helpers ───────────────────────────────────────────────────────────────────

def col_letter(n):
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def hex_to_rgb(hex_color):
    h = hex_color.lstrip("#")
    return {
        "red":   int(h[0:2], 16) / 255,
        "green": int(h[2:4], 16) / 255,
        "blue":  int(h[4:6], 16) / 255,
    }


def color_obj(hex_color):
    return {"rgbColor": hex_to_rgb(hex_color)}


def cell_range(sheet_id, start_row, end_row, start_col=0, end_col=None):
    r = {
        "sheetId":          sheet_id,
        "startRowIndex":    start_row,
        "endRowIndex":      end_row,
        "startColumnIndex": start_col,
    }
    if end_col is not None:
        r["endColumnIndex"] = end_col
    return r


def sheet_to_values(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(v) if v is not None else "" for v in row])
    return rows


def get_xl_col_widths(ws):
    """Читает ширины колонок из openpyxl (в символах)."""
    widths = {}
    for col_letter_key, dim in ws.column_dimensions.items():
        if dim.width:
            col_idx = openpyxl.utils.column_index_from_string(col_letter_key) - 1
            widths[col_idx] = dim.width
    return widths


def get_xl_fill(cell):
    """Возвращает hex цвет заливки ячейки или None."""
    try:
        fill = cell.fill
        if fill and fill.fill_type not in (None, "none"):
            fg = fill.fgColor
            if fg.type == "rgb" and fg.rgb not in ("00000000", "FFFFFFFF", "00FFFFFF"):
                return fg.rgb[-6:]  # убираем alpha
    except Exception:
        pass
    return None


def get_xl_font(cell):
    """Возвращает (bold, italic, size, color_hex) ячейки."""
    try:
        f = cell.font
        color = None
        if f.color and f.color.type == "rgb":
            c = f.color.rgb[-6:]
            if c not in ("000000", "FFFFFF"):
                color = c
        return f.bold, f.italic, f.size, color
    except Exception:
        return False, False, None, None


# ── Форматирование одного листа ───────────────────────────────────────────────

def build_format_requests(ws_xl, gsheet_id, values):
    """Генерирует batchUpdate requests для одного листа."""
    requests = []
    n_rows = len(values)
    n_cols = max((len(r) for r in values), default=1)

    # 1. Ширины колонок (из Excel)
    widths = get_xl_col_widths(ws_xl)
    for col_idx, width_chars in widths.items():
        # Excel символы → пиксели: примерно * 7 + 10
        pixels = max(60, int(width_chars * 7 + 10))
        requests.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId":    gsheet_id,
                    "dimension":  "COLUMNS",
                    "startIndex": col_idx,
                    "endIndex":   col_idx + 1,
                },
                "properties": {"pixelSize": pixels},
                "fields": "pixelSize",
            }
        })

    # 2. Перебираем все ячейки и копируем форматирование из Excel
    # Сначала читаем весь лист (не read_only — нужны стили)
    # ws_xl уже загружен с полными данными (не read_only), передаётся снаружи

    cell_formats = {}  # (row, col) → format dict

    for row_cells in ws_xl.iter_rows():
        for cell in row_cells:
            r = cell.row - 1    # 0-based
            c = cell.column - 1 # 0-based
            if r >= n_rows or c >= n_cols:
                continue

            fmt = {}

            # Заливка
            fill_hex = get_xl_fill(cell)
            if fill_hex:
                fmt["backgroundColor"] = hex_to_rgb(fill_hex)

            # Шрифт
            bold, italic, size, font_color = get_xl_font(cell)
            font_fmt = {}
            if bold:
                font_fmt["bold"] = True
            if italic:
                font_fmt["italic"] = True
            if size:
                font_fmt["fontSize"] = int(size)
            if font_color:
                font_fmt["foregroundColor"] = hex_to_rgb(font_color)
            if font_fmt:
                fmt["textFormat"] = font_fmt

            # Выравнивание
            try:
                h = cell.alignment.horizontal
                v = cell.alignment.vertical
                wrap = cell.alignment.wrap_text
                align = {}
                if h in ("center", "right", "left"):
                    align["horizontalAlignment"] = h.upper()
                if v in ("center", "top", "bottom"):
                    align["verticalAlignment"] = v.upper()
                if wrap:
                    align["wrapStrategy"] = "WRAP"
                if align:
                    fmt.update(align)
            except Exception:
                pass

            if fmt:
                cell_formats[(r, c)] = fmt

    # Группируем одинаковые форматы в диапазоны по строкам
    # Для упрощения — применяем построчно
    row_formats = {}
    for (r, c), fmt in cell_formats.items():
        row_formats.setdefault(r, {})[c] = fmt

    for row_idx, col_fmts in row_formats.items():
        for col_idx, fmt in col_fmts.items():
            fields = []
            user_fmt = {}

            if "backgroundColor" in fmt:
                user_fmt["backgroundColor"] = fmt["backgroundColor"]
                fields.append("backgroundColor")
            if "textFormat" in fmt:
                user_fmt["textFormat"] = fmt["textFormat"]
                fields.append("textFormat")
            if "horizontalAlignment" in fmt:
                user_fmt["horizontalAlignment"] = fmt["horizontalAlignment"]
                fields.append("horizontalAlignment")
            if "verticalAlignment" in fmt:
                user_fmt["verticalAlignment"] = fmt["verticalAlignment"]
                fields.append("verticalAlignment")
            if "wrapStrategy" in fmt:
                user_fmt["wrapStrategy"] = fmt["wrapStrategy"]
                fields.append("wrapStrategy")

            if not fields:
                continue

            requests.append({
                "repeatCell": {
                    "range": cell_range(gsheet_id, row_idx, row_idx + 1,
                                        col_idx, col_idx + 1),
                    "cell": {"userEnteredFormat": user_fmt},
                    "fields": "userEnteredFormat(" + ",".join(fields) + ")",
                }
            })

    # 3. Заморозить первые строки с заголовками (строки 0..2 — заголовок + подзаголовок)
    requests.append({
        "updateSheetProperties": {
            "properties": {
                "sheetId": gsheet_id,
                "gridProperties": {"frozenRowCount": 4},
            },
            "fields": "gridProperties.frozenRowCount",
        }
    })

    # 4. Высота строк — немного больше стандартной для читаемости
    requests.append({
        "updateDimensionProperties": {
            "range": {
                "sheetId":    gsheet_id,
                "dimension":  "ROWS",
                "startIndex": 0,
                "endIndex":   n_rows,
            },
            "properties": {"pixelSize": 22},
            "fields": "pixelSize",
        }
    })

    return requests


# ── Основная функция ──────────────────────────────────────────────────────────

def upload(xlsx_path, spreadsheet_id=None):
    import gspread

    print("Авторизуемся в Google…")
    creds = get_creds()
    gc = gspread.authorize(creds)

    # Загружаем Excel с полными стилями (не read_only)
    print("Читаем Excel-файл…")
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    sheet_names = wb.sheetnames

    if spreadsheet_id:
        print(f"Открываем существующую таблицу: {spreadsheet_id}")
        spreadsheet = gc.open_by_key(spreadsheet_id)
    else:
        title = "HOMEART Dashboard"
        print(f"Создаём новую таблицу: «{title}»")
        spreadsheet = gc.create(title)
        spreadsheet.share(None, perm_type="anyone", role="reader")
        print(f"  Ссылка: {spreadsheet.url}")

    existing = {ws.title: ws for ws in spreadsheet.worksheets()}

    for idx, name in enumerate(sheet_names):
        print(f"  Загружаем лист «{name}»…")
        ws_xl = wb[name]
        values = sheet_to_values(ws_xl)

        # Создаём / находим лист в Google Sheets
        if name in existing:
            gws = existing[name]
            gws.clear()
        else:
            if idx == 0 and len(spreadsheet.worksheets()) == 1:
                gws = spreadsheet.get_worksheet(0)
                gws.update_title(name)
            else:
                gws = spreadsheet.add_worksheet(
                    name,
                    rows=max(500, len(values) + 10),
                    cols=max(30, (max(len(r) for r in values) if values else 10) + 2),
                )

        # Записываем данные
        if values:
            end_col = col_letter(max(len(r) for r in values))
            gws.update(f"A1:{end_col}{len(values)}", values,
                       value_input_option="USER_ENTERED")

        # Применяем форматирование
        fmt_requests = build_format_requests(ws_xl, gws.id, values)
        if fmt_requests:
            # Разбиваем на чанки по 200 запросов (лимит API)
            for i in range(0, len(fmt_requests), 200):
                spreadsheet.batch_update({"requests": fmt_requests[i:i+200]})

        print(f"    ✓ {len(values)} строк, {len(fmt_requests)} запросов форматирования")

    wb.close()
    print(f"\n✅ Готово! Открывай: {spreadsheet.url}")
    print(f"   ID для обновления: {spreadsheet.id}")
    return spreadsheet.url, spreadsheet.id


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "dashboard_output.xlsx"
    sid  = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(xlsx):
        print(f"❌ Файл не найден: {xlsx}")
        sys.exit(1)

    upload(xlsx, sid)
