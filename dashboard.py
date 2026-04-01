#!/usr/bin/env python3
"""
Дашборд конверсии менеджеров HOMEART
Генерирует аналитический Excel на основе выгрузки из AmoCRM.

Использование:
    python3 dashboard.py "путь_к_выгрузке.xlsx"
    python3 dashboard.py  # использует первый .xlsx в текущей папке
"""

import sys
import os
from datetime import datetime
from collections import defaultdict, OrderedDict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# === КОНФИГУРАЦИЯ ВОРОНКИ ===

STAGE_ORDER = OrderedDict([
    ("новая заявка, квалификация", 1),
    ("новая заявка", 1),
    ("сделка на стопе", 0),  # пауза, не прогресс
    ("встреча, получение тз", 2),
    ("подготовка кп", 3),
    ("подготовка кп/сп", 3),
    ("отправка и согласование кп клиентом", 4),
    ("подготовка итог сп", 5),
    ("отправка и согласование итог сп", 6),
    ("ожидание предоплаты", 7),
    ("получение предоплаты", 8),
    ("атехническая (поставщики более 12)", 9),
    ("предоплата премии дизайнеру", 10),
    ("запрос и проверка проформы/счета", 11),
    ("оплата проформы", 12),
    ("постоплата премии дизайнеру", 13),
    ("поставка", 14),
    ("заказ в работе у фабрики", 15),
    ("отгрузка с фабрики/поставка на склад", 16),
    ("отгрузка на склад", 16),
    ("поступление на склад/проверка груза", 17),
    ("заявка логисту", 18),
    ("хранение на складе(доставка отложена)", 19),
    ("хранение на складе", 19),
    ("отгрузка/доставка", 20),
    ("частичная доставка", 20),
    ("полная доставка на объект", 21),
    ("отгрузка по сп завершена, акты внесены", 22),
    ("отгрузка завершена", 22),
    ("замечание", 23),
    ("завершение сделки", 90),
    ("успешно реализовано", 90),
])

RANK_LOST = -10
RANK_WON = 90
CONVERSION_STAGE_PREPAY = 8
CONVERSION_STAGE_PROFORMA = 12

SALES_FUNNEL_STAGES = [
    "новая заявка, квалификация",
    "сделка на стопе",
    "встреча, получение тз",
    "подготовка кп",
    "отправка и согласование кп клиентом",
    "подготовка итог сп",
    "отправка и согласование итог сп",
    "ожидание предоплаты",
    "получение предоплаты",
    "атехническая (поставщики более 12)",
    "предоплата премии дизайнеру",
    "запрос и проверка проформы/счета",
    "оплата проформы",
    "успешно реализовано",
    "закрыто и не реализовано",
]

STAGE_DISPLAY = {
    "новая заявка, квалификация": "Новая заявка / Квалификация",
    "новая заявка": "Новая заявка",
    "сделка на стопе": "Сделка на стопе",
    "встреча, получение тз": "Встреча / получение ТЗ",
    "подготовка кп": "Подготовка КП",
    "подготовка кп/сп": "Подготовка КП/СП",
    "отправка и согласование кп клиентом": "Отправка и согласование КП",
    "подготовка итог сп": "Подготовка итог. СП",
    "отправка и согласование итог сп": "Отправка и согласование итог. СП",
    "ожидание предоплаты": "Ожидание предоплаты",
    "получение предоплаты": "Получение предоплаты",
    "атехническая (поставщики более 12)": "Атехническая (>12 поставщиков)",
    "предоплата премии дизайнеру": "Предоплата премии дизайнеру",
    "запрос и проверка проформы/счета": "Запрос и проверка проформы",
    "оплата проформы": "Оплата проформы",
    "постоплата премии дизайнеру": "Постоплата премии дизайнеру",
    "поставка": "Поставка",
    "заказ в работе у фабрики": "Заказ в работе у фабрики",
    "отгрузка с фабрики/поставка на склад": "Отгрузка с фабрики",
    "отгрузка на склад": "Отгрузка на склад",
    "поступление на склад/проверка груза": "Поступление на склад",
    "заявка логисту": "Заявка логисту",
    "хранение на складе(доставка отложена)": "Хранение на складе",
    "хранение на складе": "Хранение на складе",
    "отгрузка/доставка": "Отгрузка / доставка",
    "частичная доставка": "Частичная доставка",
    "полная доставка на объект": "Полная доставка",
    "отгрузка по сп завершена, акты внесены": "Отгрузка завершена, акты",
    "отгрузка завершена": "Отгрузка завершена",
    "замечание": "Замечание",
    "завершение сделки": "Завершение сделки",
    "успешно реализовано": "Успешно реализовано",
    "закрыто и не реализовано": "Закрыто и не реализовано",
}


# === СТИЛИ ===

HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="Arial", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Arial", bold=True, size=11, color="2F5496")

DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGNMENT = Alignment(vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

HIGHLIGHT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


# === ПАРСИНГ ДАННЫХ ===

def parse_date(val):
    if not val or str(val).strip() in ("None", "", "не закрыта"):
        return None
    s = str(val).strip()
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_number(val):
    if not val or str(val).strip() in ("None", "", "0"):
        return 0
    s = str(val).strip().replace(" ", "").replace("\xa0", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0


def normalize_stage(stage):
    if not stage:
        return ""
    s = str(stage).strip().lower()
    if s.startswith("закрыто и не реализовано"):
        return "закрыто и не реализовано"
    return s


def stage_rank(stage_normalized):
    if stage_normalized == "закрыто и не реализовано":
        return RANK_LOST
    return STAGE_ORDER.get(stage_normalized, -1)


def load_data(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        print("Файл пуст или содержит только заголовки.")
        sys.exit(1)

    deals = []
    for row in rows[1:]:
        def col(idx):
            return row[idx] if idx < len(row) else None

        stage_raw = col(6)
        stage_norm = normalize_stage(stage_raw)
        rank = stage_rank(stage_norm)

        created = parse_date(col(9))
        prepay_date = parse_date(col(34))
        prepay_sum = parse_number(col(35))
        budget = parse_number(col(8))
        proforma_date = parse_date(col(83))  # Оплата проформы 1

        cycle_prepay = None
        if created and prepay_date and prepay_date >= created:
            cycle_prepay = (prepay_date - created).days

        cycle_proforma = None
        if created and proforma_date and proforma_date >= created:
            cycle_proforma = (proforma_date - created).days

        manager = str(col(5)).strip() if col(5) else "(без ответственного)"
        if manager in ("None", ""):
            manager = "(без ответственного)"

            # Вид товара (col 23) — может быть несколько через запятую
        raw_category = str(col(23)).strip() if col(23) else ""
        if raw_category in ("None", ""):
            raw_category = ""

        deals.append({
            "id": col(0),
            "name": col(1),
            "manager": manager,
            "stage_raw": stage_raw,
            "stage": stage_norm,
            "rank": rank,
            "budget": budget,
            "created": created,
            "close_date": parse_date(col(15)),
            "prepay_date": prepay_date,
            "prepay_sum": prepay_sum,
            "proforma_date": proforma_date,
            "postpay_date": parse_date(col(36)),
            "postpay_sum": parse_number(col(37)),
            "cycle_prepay": cycle_prepay,
            "cycle_proforma": cycle_proforma,
            "category_raw": raw_category,
        })

    return deals


# === РАСЧЁТ МЕТРИК ===

def calc_metrics(deals):
    by_manager = defaultdict(list)
    for d in deals:
        by_manager[d["manager"]].append(d)

    def metrics_for(deal_list):
        total = len(deal_list)
        lost = sum(1 for d in deal_list if d["rank"] == RANK_LOST)
        won = sum(1 for d in deal_list if d["rank"] == RANK_WON)
        active = total - lost - won

        # Получена предоплата: поле «Дата предоплаты» заполнено
        reached_prepay = sum(1 for d in deal_list if d["prepay_date"] is not None)
        # Дошли до оплаты проформы: rank >= 12 (включая завершённые с rank 90)
        reached_proforma = sum(1 for d in deal_list if d["rank"] >= CONVERSION_STAGE_PROFORMA)

        total_budget = sum(d["budget"] for d in deal_list)
        total_prepay = sum(d["prepay_sum"] for d in deal_list if d["prepay_date"] is not None)
        total_postpay = sum(d["postpay_sum"] for d in deal_list if d["postpay_sum"])
        total_inflows = total_prepay + total_postpay
        avg_budget = total_budget / total if total else 0

        # Бюджет успешных = сделки, прошедшие оплату проформы (rank >= 12)
        success_budget = sum(d["budget"] for d in deal_list if d["rank"] >= CONVERSION_STAGE_PROFORMA)

        # Средний цикл до предоплаты
        cycles_prepay = [d["cycle_prepay"] for d in deal_list if d["cycle_prepay"] is not None]
        avg_cycle_prepay = sum(cycles_prepay) / len(cycles_prepay) if cycles_prepay else None

        # Средний цикл до оплаты проформы
        cycles_proforma = [d["cycle_proforma"] for d in deal_list if d["cycle_proforma"] is not None]
        avg_cycle_proforma = sum(cycles_proforma) / len(cycles_proforma) if cycles_proforma else None

        stage_counts = defaultdict(int)
        for d in deal_list:
            stage_counts[d["stage"]] += 1

        stage_budgets = defaultdict(float)
        for d in deal_list:
            stage_budgets[d["stage"]] += d["budget"]

        return {
            "total": total,
            "lost": lost,
            "won": won,
            "active": active,
            "reached_prepay": reached_prepay,
            "reached_proforma": reached_proforma,
            "conv_prepay": (reached_prepay / total * 100) if total else 0,
            "conv_proforma": (reached_proforma / total * 100) if total else 0,
            "total_budget": total_budget,
            "success_budget": success_budget,
            "total_prepay": total_prepay,
            "total_postpay": total_postpay,
            "total_inflows": total_inflows,
            "avg_budget": avg_budget,
            "avg_cycle_prepay": avg_cycle_prepay,
            "avg_cycle_proforma": avg_cycle_proforma,
            "stage_counts": dict(stage_counts),
            "stage_budgets": dict(stage_budgets),
        }

    dept = metrics_for(deals)
    managers = {}
    for mgr, mgr_deals in sorted(by_manager.items(), key=lambda x: -len(x[1])):
        managers[mgr] = metrics_for(mgr_deals)

    return dept, managers


# === УТИЛИТЫ EXCEL ===

def apply_header_style(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def apply_data_style(ws, row, col_count, is_alt=False):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER
        if is_alt:
            cell.fill = ALT_ROW_FILL


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def fmt_pct(val):
    return f"{val:.1f}%"


def fmt_money(val):
    return f"{val:,.0f}".replace(",", " ")


def fmt_cycle(val):
    return f"{val:.0f}" if val is not None else "—"


# === ЛИСТ 1: СВОД ПО ОТДЕЛУ ===

def build_sheet_summary(wb, dept, managers, deals):
    ws = wb.create_sheet("Свод по отделу")

    ws.cell(row=1, column=1, value="Сводная аналитика отдела проектов HOMEART").font = TITLE_FONT
    ws.merge_cells("A1:E1")

    ws.cell(row=2, column=1,
            value="Конверсия рассчитана по всем сделкам, включая закрытые и успешные. "
                  "Подробности расчётов — на листе «Методология»."
            ).font = Font(name="Arial", italic=True, size=9, color="4472C4")

    r = 3
    ws.cell(row=r, column=1, value="Показатель").font = SUBTITLE_FONT
    ws.cell(row=r, column=2, value="Значение").font = SUBTITLE_FONT
    apply_header_style(ws, r, 2)

    metrics = [
        ("Всего сделок", dept["total"]),
        ("Закрыто (не реализовано)", dept["lost"]),
        ("Успешно реализовано / завершено", dept["won"]),
        ("В работе (активные)", dept["active"]),
        ("", ""),
        ("Дошло до получения предоплаты", dept["reached_prepay"]),
        ("Конверсия в предоплату", fmt_pct(dept["conv_prepay"])),
        ("Дошло до оплаты проформы", dept["reached_proforma"]),
        ("Конверсия в оплату проформы", fmt_pct(dept["conv_proforma"])),
        ("", ""),
        ("Общий бюджет всех сделок", fmt_money(dept["total_budget"]) + " ₽"),
        ("Бюджет успешных (прошли оплату проформы)", fmt_money(dept["success_budget"]) + " ₽"),
        ("Сумма предоплат", fmt_money(dept["total_prepay"]) + " ₽"),
        ("Средний чек (по всем)", fmt_money(dept["avg_budget"]) + " ₽"),
        ("", ""),
        ("Ср. цикл: создание -> предоплата",
         f"{dept['avg_cycle_prepay']:.0f} дней" if dept["avg_cycle_prepay"] else "нет данных"),
        ("Ср. цикл: создание -> оплата проформы",
         f"{dept['avg_cycle_proforma']:.0f} дней" if dept["avg_cycle_proforma"] else "нет данных"),
        ("Менеджеров в отделе", len(managers)),
    ]

    for i, (label, val) in enumerate(metrics):
        row = r + 1 + i
        ws.cell(row=row, column=1, value=label).font = DATA_FONT
        ws.cell(row=row, column=2, value=val).font = DATA_FONT
        apply_data_style(ws, row, 2, is_alt=(i % 2 == 1))

    # Воронка по отделу
    r2 = r + len(metrics) + 3
    ws.cell(row=r2, column=1, value="Воронка по отделу (текущие этапы)").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r2}:D{r2}")

    r2 += 1
    headers = ["Этап", "Сделок", "Бюджет, ₽", "% от общего"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=r2, column=c, value=h)
    apply_header_style(ws, r2, len(headers))

    for i, stage in enumerate(SALES_FUNNEL_STAGES):
        row = r2 + 1 + i
        count = dept["stage_counts"].get(stage, 0)
        budget = dept["stage_budgets"].get(stage, 0)
        pct = (count / dept["total"] * 100) if dept["total"] else 0

        ws.cell(row=row, column=1, value=STAGE_DISPLAY.get(stage, stage))
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=round(budget))
        ws.cell(row=row, column=4, value=fmt_pct(pct))
        apply_data_style(ws, row, len(headers), is_alt=(i % 2 == 1))

        if stage in ("получение предоплаты", "оплата проформы"):
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = HIGHLIGHT_FILL

    set_col_widths(ws, [45, 16, 22, 14])


# === ЛИСТ 2: КОНВЕРСИЯ ПО МЕНЕДЖЕРАМ ===

def build_sheet_conversion(wb, dept, managers):
    ws = wb.create_sheet("Конверсия по менеджерам")

    ws.cell(row=1, column=1, value="Конверсия по менеджерам").font = TITLE_FONT
    ws.merge_cells("A1:O1")

    ws.cell(row=2, column=1,
            value="Подробности расчётов каждого показателя — на листе «Методология»."
            ).font = Font(name="Arial", italic=True, size=9, color="4472C4")

    headers = [
        "Менеджер",               # 1
        "Всего\nсделок",          # 2
        "Не реали-\nзовано",      # 3
        "В работе",               # 4
        "Завершено\n(успешно)",   # 5
        "Дошло до\nпредоплаты",   # 6
        "Конверсия\n(предоплата)",# 7
        "Дошло до\nоплаты проф.", # 8
        "Конверсия\n(проформа)",  # 9
        "Общий\nбюджет, ₽",      # 10
        "Бюджет\nуспешных, ₽",   # 11
        "Сумма\nпредоплат, ₽",   # 12
        "Средний\nчек, ₽",       # 13
        "Ср. цикл до\nпредоплаты\n(дней)",    # 14
        "Ср. цикл до\nоплаты проф.\n(дней)",  # 15
    ]

    r = 3
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    apply_header_style(ws, r, len(headers))

    for i, (mgr, m) in enumerate(managers.items()):
        row = r + 1 + i
        ws.cell(row=row, column=1, value=mgr)
        ws.cell(row=row, column=2, value=m["total"])
        ws.cell(row=row, column=3, value=m["lost"])
        ws.cell(row=row, column=4, value=m["active"])
        ws.cell(row=row, column=5, value=m["won"])
        ws.cell(row=row, column=6, value=m["reached_prepay"])
        ws.cell(row=row, column=7, value=fmt_pct(m["conv_prepay"]))
        ws.cell(row=row, column=8, value=m["reached_proforma"])
        ws.cell(row=row, column=9, value=fmt_pct(m["conv_proforma"]))
        ws.cell(row=row, column=10, value=round(m["total_budget"]))
        ws.cell(row=row, column=11, value=round(m["success_budget"]))
        ws.cell(row=row, column=12, value=round(m["total_prepay"]))
        ws.cell(row=row, column=13, value=round(m["avg_budget"]))
        ws.cell(row=row, column=14, value=fmt_cycle(m["avg_cycle_prepay"]))
        ws.cell(row=row, column=15, value=fmt_cycle(m["avg_cycle_proforma"]))
        apply_data_style(ws, row, len(headers), is_alt=(i % 2 == 1))

    # Итого
    tr = r + 1 + len(managers)
    ws.cell(row=tr, column=1, value="ИТОГО")
    ws.cell(row=tr, column=2, value=dept["total"])
    ws.cell(row=tr, column=3, value=dept["lost"])
    ws.cell(row=tr, column=4, value=dept["active"])
    ws.cell(row=tr, column=5, value=dept["won"])
    ws.cell(row=tr, column=6, value=dept["reached_prepay"])
    ws.cell(row=tr, column=7, value=fmt_pct(dept["conv_prepay"]))
    ws.cell(row=tr, column=8, value=dept["reached_proforma"])
    ws.cell(row=tr, column=9, value=fmt_pct(dept["conv_proforma"]))
    ws.cell(row=tr, column=10, value=round(dept["total_budget"]))
    ws.cell(row=tr, column=11, value=round(dept["success_budget"]))
    ws.cell(row=tr, column=12, value=round(dept["total_prepay"]))
    ws.cell(row=tr, column=13, value=round(dept["avg_budget"]))
    ws.cell(row=tr, column=14, value=fmt_cycle(dept["avg_cycle_prepay"]))
    ws.cell(row=tr, column=15, value=fmt_cycle(dept["avg_cycle_proforma"]))
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=tr, column=c)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.border = THIN_BORDER
        cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    set_col_widths(ws, [36, 10, 12, 12, 12, 14, 14, 14, 14, 18, 18, 18, 16, 16, 16])


# === ЛИСТ 3: ВОРОНКА ПО МЕНЕДЖЕРАМ ===

def build_sheet_funnel(wb, managers):
    ws = wb.create_sheet("Воронка по менеджерам")

    ws.cell(row=1, column=1, value="Распределение сделок по этапам и менеджерам").font = TITLE_FONT
    ws.merge_cells(f"A1:{get_column_letter(len(SALES_FUNNEL_STAGES) + 1)}1")

    r = 3
    ws.cell(row=r, column=1, value="Менеджер")
    for j, stage in enumerate(SALES_FUNNEL_STAGES):
        ws.cell(row=r, column=2 + j, value=STAGE_DISPLAY.get(stage, stage))
    col_count = 1 + len(SALES_FUNNEL_STAGES)
    apply_header_style(ws, r, col_count)

    for i, (mgr, m) in enumerate(managers.items()):
        row = r + 1 + i
        ws.cell(row=row, column=1, value=mgr)
        for j, stage in enumerate(SALES_FUNNEL_STAGES):
            val = m["stage_counts"].get(stage, 0)
            cell = ws.cell(row=row, column=2 + j, value=val if val else "")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        apply_data_style(ws, row, col_count, is_alt=(i % 2 == 1))
        for j in range(len(SALES_FUNNEL_STAGES)):
            ws.cell(row=row, column=2 + j).alignment = Alignment(horizontal="center")

    widths = [36] + [14] * len(SALES_FUNNEL_STAGES)
    set_col_widths(ws, widths)


# === ЛИСТ 4: ВСЕ СДЕЛКИ ===

def build_sheet_deals(wb, deals):
    ws = wb.create_sheet("Все сделки")

    ws.cell(row=1, column=1, value="Все сделки — детализация").font = TITLE_FONT
    ws.merge_cells("A1:L1")

    headers = [
        "ID", "Название сделки", "Менеджер", "Этап", "Статус",
        "Бюджет, ₽", "Дата создания", "Дата предоплаты",
        "Сумма предоплаты, ₽", "Дата оплаты проформы",
        "Цикл до предоплаты (дней)", "Цикл до проформы (дней)",
        "Дата закрытия",
    ]

    r = 3
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    apply_header_style(ws, r, len(headers))

    LOST_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    WON_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    for i, d in enumerate(deals):
        row = r + 1 + i

        if d["rank"] == RANK_LOST:
            status = "Не реализовано"
        elif d["rank"] == RANK_WON:
            status = "Успешно"
        else:
            status = "В работе"

        ws.cell(row=row, column=1, value=d["id"])
        ws.cell(row=row, column=2, value=d["name"])
        ws.cell(row=row, column=3, value=d["manager"])
        ws.cell(row=row, column=4, value=STAGE_DISPLAY.get(d["stage"], d["stage_raw"]))
        ws.cell(row=row, column=5, value=status)
        ws.cell(row=row, column=6, value=round(d["budget"]) if d["budget"] else "")
        ws.cell(row=row, column=7,
                value=d["created"].strftime("%d.%m.%Y") if d["created"] else "")
        ws.cell(row=row, column=8,
                value=d["prepay_date"].strftime("%d.%m.%Y") if d["prepay_date"] else "")
        ws.cell(row=row, column=9,
                value=round(d["prepay_sum"]) if d["prepay_sum"] else "")
        ws.cell(row=row, column=10,
                value=d["proforma_date"].strftime("%d.%m.%Y") if d["proforma_date"] else "")
        ws.cell(row=row, column=11,
                value=d["cycle_prepay"] if d["cycle_prepay"] is not None else "")
        ws.cell(row=row, column=12,
                value=d["cycle_proforma"] if d["cycle_proforma"] is not None else "")
        ws.cell(row=row, column=13,
                value=d["close_date"].strftime("%d.%m.%Y") if d["close_date"] else "")
        apply_data_style(ws, row, len(headers), is_alt=(i % 2 == 1))

        if d["rank"] == RANK_LOST:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = LOST_FILL
        elif d["rank"] == RANK_WON:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = WON_FILL

    set_col_widths(ws, [12, 50, 36, 32, 16, 16, 14, 16, 18, 16, 16, 16, 14])
    ws.auto_filter.ref = f"A{r}:{get_column_letter(len(headers))}{r + len(deals)}"


# === ЛИСТ 5: МЕТОДОЛОГИЯ РАСЧЁТОВ ===

def build_sheet_methodology(wb):
    ws = wb.create_sheet("Методология")

    ws.cell(row=1, column=1, value="Методология расчётов дашборда").font = TITLE_FONT
    ws.merge_cells("A1:D1")

    ws.cell(row=2, column=1,
            value="Этот лист описывает логику расчёта каждого показателя в дашборде."
            ).font = Font(name="Arial", italic=True, size=9, color="4472C4")

    headers = ["Показатель", "Формула / логика расчёта", "Источник данных", "Примечание"]
    r = 4
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    apply_header_style(ws, r, len(headers))

    rows_data = [
        # --- Основные счётчики ---
        (
            "Всего сделок",
            "Количество всех строк в выгрузке (каждая строка = одна сделка)",
            "Кол-во строк в файле AmoCRM",
            "Включает активные, завершённые и не реализованные"
        ),
        (
            "Не реализовано",
            "Кол-во сделок, у которых этап начинается с «Закрыто и не реализовано»",
            "Колонка [6] Этап сделки",
            "Все варианты причин (купили у других, слишком дорого, и т.д.) объединяются в один статус"
        ),
        (
            "В работе",
            "Всего сделок − Не реализовано − Завершено (успешно)",
            "Расчётное поле",
            "Сделки, которые сейчас находятся на каком-либо активном этапе воронки"
        ),
        (
            "Завершено (успешно)",
            "Кол-во сделок с этапом «Успешно реализовано» или «Завершение сделки»",
            "Колонка [6] Этап сделки",
            "Финальные положительные статусы в CRM"
        ),

        # --- Конверсии ---
        (
            "Дошло до предоплаты",
            "Кол-во сделок, у которых заполнено поле «Дата предоплаты»",
            "Колонка [34] Дата предоплаты",
            "Считается факт получения денег, а не текущий этап воронки. "
            "Сделка на любом этапе (даже «Получение предоплаты») засчитывается, "
            "если в CRM зафиксирована дата предоплаты"
        ),
        (
            "Конверсия (предоплата)",
            "Дошло до предоплаты / Всего сделок × 100%",
            "Расчётное поле",
            "Считается от ВСЕХ сделок менеджера, включая не реализованные — "
            "это даёт реальную конверсию от входящего потока"
        ),
        (
            "Дошло до оплаты проформы",
            "Кол-во сделок, текущий этап которых >= «Оплата проформы» (ранг 12+)",
            "Колонка [6] Этап сделки",
            "Включает все этапы после оплаты проформы: заказ у фабрики, отгрузка, "
            "доставка, успешно реализовано и т.д."
        ),
        (
            "Конверсия (проформа)",
            "Дошло до оплаты проформы / Всего сделок × 100%",
            "Расчётное поле",
            "Считается от ВСЕХ сделок менеджера"
        ),

        # --- Финансовые ---
        (
            "Общий бюджет",
            "Сумма поля «Бюджет» по всем сделкам менеджера",
            "Колонка [8] Бюджет",
            "Бюджет всех сделок, включая не реализованные и активные"
        ),
        (
            "Бюджет успешных",
            "Сумма поля «Бюджет» по сделкам, дошедшим до оплаты проформы (ранг >= 12)",
            "Колонка [8] Бюджет + [6] Этап сделки",
            "Успешная сделка = прошла этап «Оплата проформы». "
            "После этого этапа проект считается реализованным"
        ),
        (
            "Сумма предоплат",
            "Сумма поля «Сумма предоплаты» по всем сделкам менеджера",
            "Колонка [35] Сумма предоплаты",
            "Фактически полученные предоплаты по данным CRM"
        ),
        (
            "Средний чек",
            "Общий бюджет / Всего сделок",
            "Расчётное поле",
            "Средний бюджет одной сделки менеджера (по всем, включая не реализованные)"
        ),

        # --- Циклы ---
        (
            "Ср. цикл до предоплаты (дней)",
            "Среднее( Дата предоплаты − Дата создания ) по сделкам, где обе даты заполнены",
            "Колонки [9] Дата создания и [34] Дата предоплаты",
            "Считается только для сделок, где есть и дата создания, и дата предоплаты. "
            "Показывает, сколько в среднем дней проходит от заведения сделки до получения предоплаты"
        ),
        (
            "Ср. цикл до оплаты проформы (дней)",
            "Среднее( Дата оплаты проформы 1 − Дата создания ) по сделкам, где обе даты заполнены",
            "Колонки [9] Дата создания и [83] Оплата проформы 1",
            "Считается только для сделок, где есть дата создания и дата первой оплаты проформы. "
            "Показывает полный цикл от первого касания до реальных денег за проект"
        ),

        # --- Воронка ---
        (
            "Воронка по этапам (Лист 1)",
            "Количество сделок, которые СЕЙЧАС находятся на каждом этапе",
            "Колонка [6] Этап сделки",
            "Это текущее распределение, а не кумулятивная воронка. "
            "Показывает, где сейчас «застряли» сделки"
        ),
        (
            "Воронка по менеджерам (Лист 3)",
            "Кросс-таблица: менеджер × этап = кол-во сделок на этом этапе",
            "Колонки [5] Ответственный и [6] Этап сделки",
            "Позволяет увидеть, на каких этапах у конкретного менеджера скапливаются сделки"
        ),

        # --- Статусы ---
        (
            "Статус сделки (Лист 4)",
            "«Не реализовано» если этап начинается с «Закрыто и не реализовано»; "
            "«Успешно» если этап = «Успешно реализовано» или «Завершение сделки»; "
            "«В работе» — все остальные",
            "Колонка [6] Этап сделки",
            "Упрощённый статус для удобства фильтрации"
        ),
    ]

    for i, (name, formula, source, note) in enumerate(rows_data):
        row = r + 1 + i
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=3, value=source)
        ws.cell(row=row, column=4, value=note)
        apply_data_style(ws, row, len(headers), is_alt=(i % 2 == 1))

    # Отдельный блок: порядок этапов воронки
    r3 = r + len(rows_data) + 3
    ws.cell(row=r3, column=1, value="Порядок этапов воронки (ранг)").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r3}:D{r3}")

    r3 += 1
    stage_headers = ["Этап", "Ранг", "Описание"]
    for c, h in enumerate(stage_headers, 1):
        ws.cell(row=r3, column=c, value=h)
    apply_header_style(ws, r3, len(stage_headers))

    stage_descriptions = [
        ("Сделка на стопе", "0", "Клиент на паузе, не считается прогрессом"),
        ("Новая заявка / Квалификация", "1", "Входящая заявка, первичная квалификация"),
        ("Встреча / получение ТЗ", "2", "Получение технического задания от клиента"),
        ("Подготовка КП", "3", "Менеджер готовит коммерческое предложение"),
        ("Отправка и согласование КП", "4", "КП отправлено клиенту, идёт согласование"),
        ("Подготовка итог. СП", "5", "Подготовка итоговой спецификации"),
        ("Отправка и согласование итог. СП", "6", "Итоговая СП на согласовании у клиента"),
        ("Ожидание предоплаты", "7", "Документы подписаны, ждём предоплату"),
        ("Получение предоплаты", "8", "Предоплата получена (КОНВЕРСИЯ 1)"),
        ("Атехническая (>12 поставщиков)", "9", "Техническая работа с поставщиками"),
        ("Предоплата премии дизайнеру", "10", "Внутренний процесс"),
        ("Запрос и проверка проформы", "11", "Запрос проформы у поставщиков"),
        ("Оплата проформы", "12", "Проформа оплачена (КОНВЕРСИЯ 2) — проект реализован"),
        ("Заказ у фабрики и далее", "13-90", "Дальнейшие этапы: производство, доставка, завершение"),
        ("Закрыто и не реализовано", "-10", "Сделка проиграна (с указанием причины)"),
    ]

    for i, (stage, rank, desc) in enumerate(stage_descriptions):
        row = r3 + 1 + i
        ws.cell(row=row, column=1, value=stage)
        ws.cell(row=row, column=2, value=rank)
        ws.cell(row=row, column=3, value=desc)
        apply_data_style(ws, row, len(stage_headers), is_alt=(i % 2 == 1))

        if "КОНВЕРСИЯ" in desc:
            for c in range(1, len(stage_headers) + 1):
                ws.cell(row=row, column=c).fill = HIGHLIGHT_FILL

    set_col_widths(ws, [40, 55, 35, 60])

    # --- Блок: методология листа «По категориям» ---
    r4 = r3 + len(stage_descriptions) + 4
    ws.cell(row=r4, column=1, value="Лист «По категориям» — методология расчётов").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r4}:D{r4}")

    r4 += 1
    cat_headers = ["Показатель", "Формула / логика расчёта", "Источник данных", "Примечание"]
    for c, h in enumerate(cat_headers, 1):
        ws.cell(row=r4, column=c, value=h)
    apply_header_style(ws, r4, len(cat_headers))

    cat_rows = [
        (
            "КП (кол-во)",
            "Кол-во сделок категории, дошедших до этапа «Подготовка КП» и выше (ранг ≥ 3)",
            "Колонки [6] Этап сделки, [23] Вид товара",
            "Включает все этапы начиная с «Подготовка КП»: согласование, предоплата, "
            "проформа, реализовано и т.д. Сделки на «Новая заявка» и «Встреча/ТЗ» не считаются"
        ),
        (
            "Оплачено (кол-во)",
            "Кол-во сделок, у которых заполнено поле «Дата предоплаты»",
            "Колонка [34] Дата предоплаты",
            "Считается факт получения денег, а не этап воронки. "
            "Сделка на этапе «Получение предоплаты» засчитывается, если дата в CRM заполнена"
        ),
        (
            "Отказов (закрыто)",
            "Кол-во сделок с этапом «Закрыто и не реализовано» (любая причина)",
            "Колонка [6] Этап сделки",
            "Все варианты причин отказа (купили у других, слишком дорого и т.д.) объединены"
        ),
        (
            "Конверсия КП → оплата",
            "Оплачено / КП × 100%",
            "Расчётное поле",
            "Показывает, какая доля сделок от этапа «КП» доходит до фактической предоплаты"
        ),
        (
            "Сумма предоплат, ₽",
            "Сумма поля «Сумма предоплаты» по оплаченным сделкам категории",
            "Колонка [35] Сумма предоплаты",
            "Суммируется только по сделкам с заполненной «Датой предоплаты»"
        ),
        (
            "Сумма постоплат, ₽",
            "Сумма поля «Сумма постоплаты» по оплаченным сделкам категории",
            "Колонка [37] Сумма постоплаты",
            "Постоплата = вторая часть платежа после поставки. "
            "Если постоплата не предусмотрена схемой оплаты, поле равно 0"
        ),
        (
            "Сумма поступлений, ₽",
            "Сумма предоплат + Сумма постоплат",
            "Колонки [35] + [37]",
            "Итоговые фактически полученные деньги по категории (все платежи от клиентов)"
        ),
        (
            "Средний чек оплаченных, ₽",
            "Бюджет оплаченных сделок / Кол-во оплаченных сделок",
            "Колонки [8] Бюджет, [34] Дата предоплаты",
            "Средний бюджет одной сделки среди тех, что получили предоплату. "
            "Бюджет = итоговая стоимость проекта по CRM (не сумма фактического платежа)"
        ),
        (
            "Бюджет всех КП, ₽",
            "Сумма поля «Бюджет» по всем сделкам, дошедшим до этапа КП",
            "Колонки [8] Бюджет, [6] Этап сделки",
            "Потенциальный объём всех поданных коммерческих предложений по категории"
        ),
        (
            "Одна сделка — несколько категорий",
            "Если в поле «Вид товара» указано несколько категорий через запятую "
            "(например, «Мебель, Свет»), сделка засчитывается в каждую из них",
            "Колонка [23] Вид товара",
            "Поэтому сумма по строкам таблицы может превышать общий итог по отделу — "
            "это нормально и ожидаемо"
        ),
    ]

    for i, (name, formula, source, note) in enumerate(cat_rows):
        row = r4 + 1 + i
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=3, value=source)
        ws.cell(row=row, column=4, value=note)
        apply_data_style(ws, row, len(cat_headers), is_alt=(i % 2 == 1))


# === ЛИСТ 6: ПО КАТЕГОРИЯМ ===

# Канонические категории и ключевые слова для маппинга
CATEGORY_MAP = [
    ("Мебель",              ["мебель", "гардероб", "уличная мебель", "матрас", "ковры", "аксессуары"]),
    ("Кухни",               ["кухня", "кухн"]),
    ("Сантехника",          ["сантехника", "сантех"]),
    ("Двери",               ["двери", "дверь"]),
    ("Свет декоративный",   ["свет", "декоратив"]),
    ("Свет технический",    ["технический свет", "техсвет"]),
    ("Столярка",            ["столярка", "столяр", "индивидуальное производство"]),
]
CATEGORY_OTHER = "Другие"


def classify_categories(raw_category):
    """Возвращает список канонических категорий для одной сделки."""
    if not raw_category:
        return [CATEGORY_OTHER]

    parts = [p.strip().lower() for p in raw_category.split(",")]
    result = set()

    for part in parts:
        matched = False
        # Технический свет проверяем раньше, чтобы не попасть под «свет»
        if "технический свет" in part or "техсвет" in part:
            result.add("Свет технический")
            matched = True
        else:
            for canon, keywords in CATEGORY_MAP:
                for kw in keywords:
                    if kw in part:
                        result.add(canon)
                        matched = True
                        break
                if matched:
                    break
        if not matched:
            result.add(CATEGORY_OTHER)

    return list(result) if result else [CATEGORY_OTHER]


def calc_category_metrics(deals):
    """Считает метрики по категориям товаров."""
    all_categories = [c for c, _ in CATEGORY_MAP] + [CATEGORY_OTHER]
    data = {cat: {"kp": 0, "paid": 0, "lost": 0, "budget_all": 0,
                  "budget_paid": 0, "prepay_sum": 0, "postpay_sum": 0} for cat in all_categories}

    for d in deals:
        cats = classify_categories(d.get("category_raw", ""))
        # КП = сделки, дошедшие хотя бы до этапа «Подготовка КП» (rank >= 3)
        is_kp = d["rank"] >= 3 or d["rank"] == RANK_WON
        # Оплачено = заполнено поле «Дата предоплаты» (реальный факт получения денег)
        is_paid = d["prepay_date"] is not None
        is_lost = d["rank"] == RANK_LOST

        for cat in cats:
            if cat not in data:
                data[cat] = {"kp": 0, "paid": 0, "lost": 0, "budget_all": 0,
                             "budget_paid": 0, "prepay_sum": 0, "postpay_sum": 0}
            data[cat]["budget_all"] += d["budget"]
            if is_kp:
                data[cat]["kp"] += 1
            if is_paid:
                data[cat]["paid"] += 1
                data[cat]["budget_paid"] += d["budget"]
                data[cat]["prepay_sum"] += d["prepay_sum"]
                data[cat]["postpay_sum"] += d.get("postpay_sum") or 0
            if is_lost:
                data[cat]["lost"] += 1

    return data


def build_sheet_categories(wb, deals):
    ws = wb.create_sheet("По категориям")

    ws.cell(row=1, column=1, value="Аналитика по категориям товаров").font = TITLE_FONT
    ws.merge_cells("A1:I1")

    ws.cell(row=2, column=1,
            value="КП = сделки, дошедшие до этапа «Подготовка КП» и выше. "
                  "Оплачено = заполнено поле «Дата предоплаты» (факт получения денег). "
                  "Одна сделка может входить в несколько категорий."
            ).font = Font(name="Arial", italic=True, size=9, color="4472C4")

    headers = [
        "Категория",
        "КП\n(кол-во)",
        "Оплачено\n(кол-во)",
        "Отказов\n(закрыто)",
        "Конверсия\nКП → оплата",
        "Сумма\nпредоплат, ₽",
        "Сумма\nпостоплат, ₽",
        "Сумма\nпоступлений, ₽",
        "Средний чек\nоплаченных, ₽",
        "Бюджет\nвсех КП, ₽",
    ]

    r = 4
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    apply_header_style(ws, r, len(headers))

    cat_data = calc_category_metrics(deals)
    all_categories = [c for c, _ in CATEGORY_MAP] + [CATEGORY_OTHER]

    totals = {"kp": 0, "paid": 0, "lost": 0,
              "budget_all": 0, "budget_paid": 0, "prepay_sum": 0, "postpay_sum": 0}

    for i, cat in enumerate(all_categories):
        m = cat_data.get(cat, {})
        kp = m.get("kp", 0)
        paid = m.get("paid", 0)
        lost = m.get("lost", 0)
        budget_all = m.get("budget_all", 0)
        budget_paid = m.get("budget_paid", 0)
        prepay_sum = m.get("prepay_sum", 0)
        postpay_sum = m.get("postpay_sum", 0)
        total_inflows = prepay_sum + postpay_sum
        conv = (paid / kp * 100) if kp else 0
        avg_check = (budget_paid / paid) if paid else 0

        row = r + 1 + i
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=kp)
        ws.cell(row=row, column=3, value=paid)
        ws.cell(row=row, column=4, value=lost)
        ws.cell(row=row, column=5, value=fmt_pct(conv))
        ws.cell(row=row, column=6, value=round(prepay_sum))
        ws.cell(row=row, column=7, value=round(postpay_sum) if postpay_sum else 0)
        ws.cell(row=row, column=8, value=round(total_inflows))
        ws.cell(row=row, column=9, value=round(avg_check))
        ws.cell(row=row, column=10, value=round(budget_all))
        apply_data_style(ws, row, len(headers), is_alt=(i % 2 == 1))

        # Подсвечиваем лучшую конверсию
        if conv >= 50:
            ws.cell(row=row, column=5).fill = HIGHLIGHT_FILL

        totals["kp"] += kp
        totals["paid"] += paid
        totals["lost"] += lost
        totals["budget_all"] += budget_all
        totals["budget_paid"] += budget_paid
        totals["prepay_sum"] += prepay_sum
        totals["postpay_sum"] += postpay_sum

    # Итого (не суммируем, т.к. сделки могут входить в несколько категорий)
    note_row = r + len(all_categories) + 2
    ws.cell(row=note_row, column=1,
            value="* Сумма по строкам может превышать итог по отделу, "
                  "т.к. одна сделка с несколькими категориями считается в каждой."
            ).font = Font(name="Arial", italic=True, size=9, color="996633")

    # Второй блок: топ сделок по категории
    r2 = note_row + 3
    ws.cell(row=r2, column=1, value="Топ-5 крупнейших оплаченных сделок по категории").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r2}:F{r2}")

    r2 += 1
    top_headers = ["Категория", "Название сделки", "Менеджер", "Бюджет, ₽", "Предоплата, ₽", "Этап"]
    for c, h in enumerate(top_headers, 1):
        ws.cell(row=r2, column=c, value=h)
    apply_header_style(ws, r2, len(top_headers))

    current_row = r2 + 1
    for cat in all_categories:
        paid_deals = [
            d for d in deals
            if d["prepay_date"] is not None
            and cat in classify_categories(d.get("category_raw", ""))
        ]
        paid_deals.sort(key=lambda x: -x["budget"])
        for d in paid_deals[:5]:
            ws.cell(row=current_row, column=1, value=cat)
            ws.cell(row=current_row, column=2, value=d["name"])
            ws.cell(row=current_row, column=3, value=d["manager"])
            ws.cell(row=current_row, column=4, value=round(d["budget"]))
            ws.cell(row=current_row, column=5, value=round(d["prepay_sum"]))
            ws.cell(row=current_row, column=6, value=STAGE_DISPLAY.get(d["stage"], d["stage_raw"]))
            apply_data_style(ws, current_row, len(top_headers), is_alt=(current_row % 2 == 0))
            current_row += 1

    set_col_widths(ws, [22, 12, 12, 12, 16, 18, 18, 20, 18, 18])


# === ГЕНЕРАЦИЯ ===

def generate_dashboard(input_path, output_path):
    print(f"Загружаю данные из: {input_path}")
    deals = load_data(input_path)
    print(f"Загружено сделок: {len(deals)}")

    dept, managers = calc_metrics(deals)

    print(f"Менеджеров: {len(managers)}")
    print(f"Конверсия в предоплату (отдел): {dept['conv_prepay']:.1f}%")
    print(f"Конверсия в оплату проформы (отдел): {dept['conv_proforma']:.1f}%")
    if dept['avg_cycle_prepay']:
        print(f"Ср. цикл до предоплаты: {dept['avg_cycle_prepay']:.0f} дней")
    if dept['avg_cycle_proforma']:
        print(f"Ср. цикл до оплаты проформы: {dept['avg_cycle_proforma']:.0f} дней")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_sheet_summary(wb, dept, managers, deals)
    build_sheet_conversion(wb, dept, managers)
    build_sheet_funnel(wb, managers)
    build_sheet_categories(wb, deals)
    build_sheet_deals(wb, deals)
    build_sheet_methodology(wb)

    wb.save(output_path)
    print(f"\nДашборд сохранён: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        xlsx_files = [f for f in os.listdir(".") if f.endswith(".xlsx")
                      and not f.startswith("dashboard_") and not f.startswith("~$")]
        if not xlsx_files:
            print("Ошибка: укажи путь к выгрузке AmoCRM или положи .xlsx в текущую папку.")
            sys.exit(1)
        input_file = xlsx_files[0]
        print(f"Найден файл: {input_file}")

    if not os.path.exists(input_file):
        print(f"Ошибка: файл не найден: {input_file}")
        sys.exit(1)

    output_file = os.path.join(os.path.dirname(input_file) or ".", "dashboard_output.xlsx")
    generate_dashboard(input_file, output_file)
