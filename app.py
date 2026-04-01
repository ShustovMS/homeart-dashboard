#!/usr/bin/env python3
"""
HOMEART Dashboard Generator — веб-интерфейс
Запуск: python3 app.py
Открыть в браузере: http://localhost:5050
"""

import os
import sys
import uuid
import threading
import webbrowser
from pathlib import Path

# На продакшене (Railway) браузер не открываем
IS_PROD = os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("PORT")

from flask import Flask, request, send_file, render_template_string, jsonify

# Добавляем текущую папку в путь, чтобы импортировать dashboard
sys.path.insert(0, os.path.dirname(__file__))
import dashboard as db

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

UPLOAD_DIR = Path(__file__).parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# ── HTML-шаблон ──────────────────────────────────────────────────────────────

HTML = """
<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>HOMEART · Генератор дашборда</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    background: #f0f4f8;
    min-height: 100vh;
    display: flex;
    align-items: center;
    justify-content: center;
    padding: 24px;
  }
  .card {
    background: #fff;
    border-radius: 16px;
    box-shadow: 0 4px 24px rgba(0,0,0,.10);
    padding: 48px 56px;
    max-width: 560px;
    width: 100%;
  }
  .logo { font-size: 13px; font-weight: 600; letter-spacing: 2px;
          color: #2F5496; text-transform: uppercase; margin-bottom: 8px; }
  h1 { font-size: 26px; color: #1a1a2e; margin-bottom: 6px; }
  .sub { font-size: 14px; color: #6b7280; margin-bottom: 40px; line-height: 1.5; }

  .drop-zone {
    border: 2px dashed #c7d2e0;
    border-radius: 12px;
    padding: 40px 24px;
    text-align: center;
    cursor: pointer;
    transition: all .2s;
    background: #f8fafc;
  }
  .drop-zone:hover, .drop-zone.drag { border-color: #2F5496; background: #eef2fb; }
  .drop-zone input[type=file] { display: none; }
  .drop-icon { font-size: 40px; margin-bottom: 12px; }
  .drop-label { font-size: 15px; color: #374151; margin-bottom: 4px; }
  .drop-hint { font-size: 13px; color: #9ca3af; }
  .file-name { margin-top: 12px; font-size: 14px; color: #2F5496; font-weight: 500;
               display: none; }

  .btn {
    display: block;
    width: 100%;
    margin-top: 24px;
    padding: 14px;
    background: #2F5496;
    color: #fff;
    border: none;
    border-radius: 10px;
    font-size: 16px;
    font-weight: 600;
    cursor: pointer;
    transition: background .2s;
  }
  .btn:hover { background: #1e3a6e; }
  .btn:disabled { background: #9ca3af; cursor: not-allowed; }

  .progress { display: none; margin-top: 20px; }
  .progress-bar-wrap {
    background: #e5e7eb;
    border-radius: 99px;
    height: 8px;
    overflow: hidden;
    margin-bottom: 8px;
  }
  .progress-bar {
    height: 100%;
    background: #2F5496;
    border-radius: 99px;
    width: 0%;
    transition: width .3s;
    animation: indeterminate 1.4s infinite ease-in-out;
  }
  @keyframes indeterminate {
    0%   { width: 0%;   margin-left: 0; }
    50%  { width: 60%;  margin-left: 20%; }
    100% { width: 0%;   margin-left: 100%; }
  }
  .progress-label { font-size: 13px; color: #6b7280; text-align: center; }

  .result { display: none; margin-top: 24px; text-align: center; }
  .result-icon { font-size: 36px; margin-bottom: 8px; }
  .result-title { font-size: 18px; font-weight: 600; color: #16a34a; margin-bottom: 4px; }
  .result-sub { font-size: 13px; color: #6b7280; margin-bottom: 20px; }
  .dl-btn {
    display: inline-block;
    padding: 12px 32px;
    background: #16a34a;
    color: #fff;
    border-radius: 10px;
    font-size: 15px;
    font-weight: 600;
    text-decoration: none;
    transition: background .2s;
  }
  .dl-btn:hover { background: #15803d; }

  .error { display: none; margin-top: 20px; padding: 14px 16px;
           background: #fef2f2; border-radius: 10px;
           color: #dc2626; font-size: 14px; }

  .stats {
    margin-top: 24px;
    border-top: 1px solid #e5e7eb;
    padding-top: 20px;
    display: none;
  }
  .stats-title { font-size: 13px; font-weight: 600; color: #6b7280;
                 text-transform: uppercase; letter-spacing: 1px; margin-bottom: 12px; }
  .stats-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }
  .stat { background: #f8fafc; border-radius: 8px; padding: 12px 14px; }
  .stat-val { font-size: 20px; font-weight: 700; color: #1a1a2e; }
  .stat-lbl { font-size: 12px; color: #9ca3af; margin-top: 2px; }

  .period-block {
    margin-top: 20px;
    background: #f8fafc;
    border-radius: 12px;
    padding: 16px 18px;
    border: 1px solid #e5e7eb;
  }
  .period-label {
    font-size: 13px; font-weight: 600; color: #374151;
    margin-bottom: 4px;
  }
  .period-hint {
    font-size: 12px; color: #9ca3af; margin-bottom: 12px;
  }
  .period-row { display: flex; gap: 12px; align-items: center; }
  .period-row label { font-size: 13px; color: #6b7280; white-space: nowrap; }
  .period-row input[type=date] {
    flex: 1;
    padding: 8px 10px;
    border: 1px solid #d1d5db;
    border-radius: 8px;
    font-size: 14px;
    color: #1a1a2e;
    background: #fff;
  }
  .period-row input[type=date]:focus {
    outline: none; border-color: #2F5496;
  }
  .period-sep { font-size: 13px; color: #9ca3af; }
</style>
</head>
<body>
<div class="card">
  <div class="logo">HOMEART</div>
  <h1>Генератор дашборда</h1>
  <p class="sub">Загрузи выгрузку из AmoCRM (или другой CRM) в формате&nbsp;.xlsx&nbsp;—
     получишь готовый аналитический файл с 6 листами.</p>

  <form id="form" enctype="multipart/form-data">
    <div class="drop-zone" id="dropZone" onclick="document.getElementById('fileInput').click()">
      <input type="file" id="fileInput" name="file" accept=".xlsx,.csv">
      <div class="drop-icon">📂</div>
      <div class="drop-label">Нажми или перетащи файл сюда</div>
      <div class="drop-hint">.xlsx · до 50 МБ</div>
      <div class="file-name" id="fileName"></div>
    </div>
    <div class="period-block">
      <div class="period-label">Период анализа <span style="font-weight:400;color:#9ca3af">(необязательно)</span></div>
      <div class="period-hint">
        Если выгрузка за широкий период — укажи нужный квартал.<br>
        Воронка считается по дате <b>создания</b> сделки, поступления — по дате <b>предоплаты</b>.
      </div>
      <div class="period-row">
        <label>с</label>
        <input type="date" id="periodFrom" name="period_from">
        <span class="period-sep">—</span>
        <label>по</label>
        <input type="date" id="periodTo" name="period_to">
      </div>
    </div>
    <button class="btn" id="btn" type="submit" disabled>Сгенерировать дашборд</button>
  </form>

  <div class="progress" id="progress">
    <div class="progress-bar-wrap"><div class="progress-bar" id="bar"></div></div>
    <div class="progress-label" id="progressLabel">Обрабатываем файл…</div>
  </div>

  <div class="error" id="error"></div>

  <div class="result" id="result">
    <div class="result-icon">✅</div>
    <div class="result-title">Дашборд готов!</div>
    <div class="result-sub" id="resultSub"></div>
    <div style="display:flex;gap:12px;justify-content:center;flex-wrap:wrap;margin-top:4px">
      <a class="dl-btn" id="dlBtn" href="#">⬇ Скачать .xlsx</a>
      <button class="dl-btn" id="sheetsBtn"
              style="background:#1a73e8;border:none;cursor:pointer"
              onclick="publishSheets()">
        📊 Открыть в Google Sheets
      </button>
    </div>
    <div id="sheetsStatus" style="margin-top:12px;font-size:13px;color:#6b7280;display:none"></div>
  </div>

  <div class="stats" id="stats">
    <div class="stats-title">Ключевые показатели</div>
    <div class="stats-grid" id="statsGrid"></div>
  </div>
</div>

<script>
const dropZone = document.getElementById('dropZone');
const fileInput = document.getElementById('fileInput');
const btn = document.getElementById('btn');
const fileName = document.getElementById('fileName');

fileInput.addEventListener('change', () => {
  if (fileInput.files[0]) {
    fileName.textContent = '📄 ' + fileInput.files[0].name;
    fileName.style.display = 'block';
    btn.disabled = false;
  }
});

['dragover','dragenter'].forEach(e => dropZone.addEventListener(e, ev => {
  ev.preventDefault(); dropZone.classList.add('drag');
}));
['dragleave','drop'].forEach(e => dropZone.addEventListener(e, ev => {
  dropZone.classList.remove('drag');
}));
dropZone.addEventListener('drop', ev => {
  ev.preventDefault();
  fileInput.files = ev.dataTransfer.files;
  fileInput.dispatchEvent(new Event('change'));
});

let _token = null;

async function publishSheets() {
  const btn = document.getElementById('sheetsBtn');
  const status = document.getElementById('sheetsStatus');
  btn.disabled = true;
  btn.textContent = '⏳ Публикуем…';
  status.style.display = 'block';
  status.textContent = 'Загружаем данные в Google Sheets…';
  try {
    const resp = await fetch('/publish_sheets/' + _token, { method: 'POST' });
    const data = await resp.json();
    if (data.ok) {
      status.innerHTML = `✅ Готово! <a href="${data.url}" target="_blank" style="color:#1a73e8">Открыть таблицу →</a>`;
      btn.textContent = '📊 Открыть →';
      btn.onclick = () => window.open(data.url, '_blank');
    } else {
      status.textContent = '❌ ' + data.error;
      btn.textContent = '📊 Открыть в Google Sheets';
    }
  } catch(err) {
    status.textContent = '❌ Ошибка: ' + err;
    btn.textContent = '📊 Открыть в Google Sheets';
  }
  btn.disabled = false;
}

document.getElementById('form').addEventListener('submit', async (e) => {
  e.preventDefault();
  const file = fileInput.files[0];
  if (!file) return;

  btn.disabled = true;
  document.getElementById('progress').style.display = 'block';
  document.getElementById('error').style.display = 'none';
  document.getElementById('result').style.display = 'none';
  document.getElementById('stats').style.display = 'none';
  document.getElementById('progressLabel').textContent = 'Загружаем файл…';

  const fd = new FormData();
  fd.append('file', file);
  const pFrom = document.getElementById('periodFrom').value;
  const pTo   = document.getElementById('periodTo').value;
  if (pFrom) fd.append('period_from', pFrom);
  if (pTo)   fd.append('period_to',   pTo);

  try {
    const resp = await fetch('/generate', { method: 'POST', body: fd });
    const data = await resp.json();

    document.getElementById('progress').style.display = 'none';

    if (data.ok) {
      document.getElementById('resultSub').textContent =
        `${data.deals} сделок · ${data.managers} менеджеров`;
      _token = data.token;
      document.getElementById('dlBtn').href = '/download/' + data.token;
      document.getElementById('result').style.display = 'block';

      // Stats
      const grid = document.getElementById('statsGrid');
      grid.innerHTML = '';
      data.stats.forEach(s => {
        grid.innerHTML += `<div class="stat">
          <div class="stat-val">${s.val}</div>
          <div class="stat-lbl">${s.lbl}</div>
        </div>`;
      });
      document.getElementById('stats').style.display = 'block';
    } else {
      document.getElementById('error').textContent = '❌ ' + data.error;
      document.getElementById('error').style.display = 'block';
    }
  } catch (err) {
    document.getElementById('progress').style.display = 'none';
    document.getElementById('error').textContent = '❌ Ошибка соединения: ' + err;
    document.getElementById('error').style.display = 'block';
  }

  btn.disabled = false;
});
</script>
</body>
</html>
"""

# ── Маршруты ─────────────────────────────────────────────────────────────────

_files = {}  # token → output path


@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/generate", methods=["POST"])
def generate():
    from datetime import datetime as dt
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify(ok=False, error="Файл не выбран")

    token = str(uuid.uuid4())[:8]
    input_path = UPLOAD_DIR / f"{token}_input.xlsx"
    output_path = UPLOAD_DIR / f"{token}_dashboard.xlsx"

    f.save(str(input_path))

    # Период анализа (необязательный)
    def parse_date(s):
        try:
            return dt.strptime(s, "%Y-%m-%d") if s else None
        except Exception:
            return None

    period_start = parse_date(request.form.get("period_from", ""))
    period_end   = parse_date(request.form.get("period_to", ""))

    try:
        deals = db.load_data(str(input_path))
        dept, managers = db.calc_metrics(deals, period_start, period_end)

        wb = __import__("openpyxl").Workbook()
        wb.remove(wb.active)
        db.build_sheet_summary(wb, dept, managers, deals, period_start, period_end)
        db.build_sheet_conversion(wb, dept, managers, period_start, period_end)
        db.build_sheet_funnel(wb, managers, period_start, period_end)
        db.build_sheet_categories(wb, deals, period_start, period_end)
        db.build_sheet_deals(wb, deals, period_start, period_end)
        db.build_sheet_methodology(wb)
        wb.save(str(output_path))

        _files[token] = str(output_path)

        stats = [
            {"val": str(dept["total"]), "lbl": "Всего сделок"},
            {"val": db.fmt_pct(dept["conv_proforma"]), "lbl": "Конверсия (проформа)"},
            {"val": db.fmt_money(dept["total_budget"]) + " ₽", "lbl": "Общий бюджет"},
            {"val": db.fmt_cycle(dept["avg_cycle_proforma"]) + " дн.", "lbl": "Ср. цикл"},
        ]

        return jsonify(
            ok=True,
            token=token,
            deals=dept["total"],
            managers=len(managers),
            stats=stats,
        )

    except Exception as ex:
        return jsonify(ok=False, error=str(ex))

    finally:
        try:
            input_path.unlink()
        except Exception:
            pass


@app.route("/publish_sheets/<token>", methods=["POST"])
def publish_sheets(token):
    path = _files.get(token)
    if not path or not os.path.exists(path):
        return jsonify(ok=False, error="Файл не найден. Сначала сгенерируй дашборд.")
    try:
        # На сервере — service account из переменной окружения
        sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
        if sa_json:
            url, sid = _upload_via_service_account(path, sa_json)
        else:
            # Локально — OAuth
            import sheets_upload
            url, sid = sheets_upload.upload(path)
        return jsonify(ok=True, url=url, spreadsheet_id=sid)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex))


def _upload_via_service_account(xlsx_path, sa_json_str):
    """Загружает в Google Sheets через service account (для сервера)."""
    import json
    import gspread
    from google.oauth2.service_account import Credentials as SACredentials

    sa_info = json.loads(sa_json_str)
    creds = SACredentials.from_service_account_info(
        sa_info,
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ],
    )
    gc = gspread.authorize(creds)

    import openpyxl as xl
    wb = xl.load_workbook(xlsx_path, read_only=True)

    title = "HOMEART Dashboard"
    spreadsheet = gc.create(title)
    spreadsheet.share(None, perm_type="anyone", role="reader")

    existing = {ws.title: ws for ws in spreadsheet.worksheets()}

    for idx, name in enumerate(wb.sheetnames):
        ws_xl = wb[name]
        values = [[str(v) if v is not None else "" for v in row]
                  for row in ws_xl.iter_rows(values_only=True)]

        if name in existing:
            gws = existing[name]
            gws.clear()
        elif idx == 0:
            gws = spreadsheet.get_worksheet(0)
            gws.update_title(name)
        else:
            gws = spreadsheet.add_worksheet(name, rows=500, cols=30)

        if values:
            from sheets_upload import col_letter
            end_col = col_letter(max(len(r) for r in values))
            gws.update(f"A1:{end_col}{len(values)}", values)

    wb.close()
    return spreadsheet.url, spreadsheet.id


@app.route("/download/<token>")
def download(token):
    path = _files.get(token)
    if not path or not os.path.exists(path):
        return "Файл не найден или истёк срок", 404
    return send_file(path, as_attachment=True, download_name="dashboard_homeart.xlsx")


# ── Запуск ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    print(f"\n{'='*50}")
    print(f"  HOMEART Dashboard Generator")
    if not IS_PROD:
        url = f"http://localhost:{port}"
        print(f"  Открой в браузере: {url}")
        threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    print(f"{'='*50}\n")
    app.run(host="0.0.0.0", port=port, debug=False)
